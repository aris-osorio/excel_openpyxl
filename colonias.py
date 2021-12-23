from openpyxl import Workbook
from openpyxl import load_workbook


def initial_excel():
    nuevo_excel = Workbook()
    #initial headers
    nuevo_excel.worksheets[0]['A1'].value = 'id'
    nuevo_excel.worksheets[0]['B1'].value = 'postal_code'
    nuevo_excel.worksheets[0]['C1'].value = 'name'    
    return nuevo_excel

colonias_excel = load_workbook('COLONIAS_MEXICO.xlsx')
hoja = colonias_excel['Hoja1']
columnas = hoja.max_row
contador = 2
nuevo_excel = initial_excel()

while contador <= columnas:
    nuevo_excel.worksheets[0][f'A{contador}'].value = hoja[f'A{contador}'].value
    nuevo_excel.worksheets[0][f'B{contador}'].value = hoja[f'B{contador}'].value
    nuevo_excel.worksheets[0][f'C{contador}'].value = hoja[f'C{contador}'].value
    print(contador)
    contador += 1

nuevo_excel.save(f'COLONIAS_MEXICO_COPIA.xlsx')